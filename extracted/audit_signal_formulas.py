"""Audit the exact Yes/No signal formulas in EMA / Drawdown / Zone / Aggregate workbooks.

Read-only. Pulls the real formula text (incl. ArrayFormula.text) for the signal
columns so we can reconcile against signals.py / pillars.py.
"""
from __future__ import annotations
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

SHEETS_DIR = "extracted/sheets/"


def ftext(v):
    if isinstance(v, ArrayFormula):
        return "{ARRAY} " + v.text
    return v


def dump_summary(f, sheet="Summary", header_row=2, rows=(3, 4)):
    wb = load_workbook(SHEETS_DIR + f + ".xlsx", data_only=False)
    ws = wb[sheet]
    print(f"\n========== {f} / {sheet}  (dims {ws.dimensions}) ==========")
    # header row
    hdr = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None:
            hdr[c] = v
    print("HEADERS:", {get_column_letter(c): v for c, v in hdr.items()})
    for r in rows:
        print(f"-- row {r} --")
        for c in hdr:
            v = ws.cell(r, c).value
            print(f"  {get_column_letter(c)}{r} [{hdr[c]}] = {ftext(v)}")
    wb.close()


# EMA / Drawdown / Zone: per-stock Yes/No live snapshot in Summary
for f in ["Stocks_EMA_Signal", "Stocks_Drawdown_Signal", "Stocks_Zone_Signal"]:
    dump_summary(f)
