"""Extract formulas from xlsx sheets to audit for look-ahead bias."""
from __future__ import annotations
import sys
from pathlib import Path
from openpyxl import load_workbook

# Target column headers we care about (case-insensitive substring match)
TARGETS = [
    "Gain_100DMA_50DMA",
    "Gain_50DMA_50DMA",
    "Gain_200DMA_50DMA",
    "100DMA_Gain",
    "50DMA_Gain",
    "200DMA_Gain",
    "300DMA_Gain",
    "Price_Signal_50DMA",
    "Price_Signal",
    "STD_Signal",
    "Price_300DMA",
    "Price_200DMA",
    "Price_100DMA",
    "Price_50DMA",
    "Price_20DMA",
    "Price Gain",
    "Price_GT_50DMA",
    "50DMA_GT_100DMA",
    "100DMA_GT_200DMA",
    "Momentum_100DMA_GT_200DMA",
]


def normalize(s: str) -> str:
    return s.strip().replace("_", "").replace(" ", "").lower() if s else ""


def find_header_row(ws, max_rows: int = 50) -> tuple[int, dict[str, int]] | None:
    """Find a row that contains at least 3 target headers; return (row_idx, header_to_col_idx)."""
    targets_norm = {normalize(t): t for t in TARGETS}
    for row_idx in range(1, min(max_rows, ws.max_row) + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        col_to_target = {}
        for col_idx, val in enumerate(row, start=1):
            if not val or not isinstance(val, str):
                continue
            n = normalize(val)
            if n in targets_norm:
                col_to_target[col_idx] = targets_norm[n]
        if len(col_to_target) >= 3:
            return row_idx, col_to_target
    return None


def first_formula_row(ws, header_row: int, col_idx: int, max_scan: int = 400) -> tuple[int, str] | None:
    """Find the first row below header_row where col_idx contains a formula (starts with =)."""
    for r in range(header_row + 1, min(header_row + max_scan, ws.max_row) + 1):
        cell = ws.cell(row=r, column=col_idx)
        val = cell.value
        if val and isinstance(val, str) and val.startswith("="):
            return r, val
    return None


def dump_all_formulas(ws, max_rows: int = 80, max_cols: int = 30):
    """For small sheets like Summary — dump every cell that has a formula or label."""
    print(f"  --- All non-empty cells (up to row {max_rows}, col {max_cols}) ---")
    for r in range(1, min(max_rows, ws.max_row) + 1):
        for c in range(1, min(max_cols, ws.max_column) + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None or v == "":
                continue
            label = f"  [{cell.coordinate}]"
            if isinstance(v, str) and v.startswith("="):
                print(f"{label} FORMULA: {v}")
            elif isinstance(v, str) and len(v) < 200:
                print(f"{label} TEXT: {v}")
            elif isinstance(v, (int, float)):
                if abs(v) > 1e-9:
                    print(f"{label} NUM: {v}")


def extract_from_workbook(path: Path):
    print(f"\n{'=' * 80}")
    print(f"FILE: {path.name}")
    print(f"{'=' * 80}")
    wb = load_workbook(path, data_only=False, read_only=False)
    print(f"Sheets/tabs: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 5 or ws.max_column < 3:
            continue
        # Force dump for these key tabs
        if sheet_name.lower() in {"summary", "stock_list", "rules", "results"}:
            print(f"--- Sheet: '{sheet_name}' (KEY TAB, dims {ws.max_row}r × {ws.max_column}c) ---")
            dump_all_formulas(ws, max_rows=200, max_cols=20)
            print()
            continue
        # Small sheets: dump everything
        if ws.max_row <= 80 and ws.max_column <= 30:
            print(f"--- Sheet: '{sheet_name}' (small, dims {ws.max_row}r × {ws.max_column}c) ---")
            dump_all_formulas(ws)
            print()
            continue
        result = find_header_row(ws)
        if not result:
            continue
        header_row, col_to_target = result
        print(f"--- Sheet: '{sheet_name}' (header row {header_row}, dims {ws.max_row}r × {ws.max_column}c) ---")
        for col_idx, header in sorted(col_to_target.items(), key=lambda kv: TARGETS.index(kv[1]) if kv[1] in TARGETS else 999):
            r = first_formula_row(ws, header_row, col_idx)
            col_letter = ws.cell(row=header_row, column=col_idx).coordinate.rstrip("0123456789")
            if r is None:
                print(f"  [{col_letter}] {header}: <no formula in scan window>")
            else:
                row_num, formula = r
                print(f"  [{col_letter}{row_num}] {header}:")
                print(f"      {formula}")
        print()


if __name__ == "__main__":
    sheets_dir = Path(r"c:\Users\vilas\Downloads\AI_Hedge_Fund\extracted\sheets")
    targets = sys.argv[1:] if len(sys.argv) > 1 else [
        "Stocks_Buy_Signal_Analyser_15Yr.xlsx",  # PRIORITY 1
    ]
    for name in targets:
        extract_from_workbook(sheets_dir / name)
