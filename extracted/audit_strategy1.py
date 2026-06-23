"""Forensic reconciliation of Strategy 1 against Stocks_Buy_Signal_Analyser_15Yr.xlsx.

Approach:
1. Read all cached values from the xlsx Stock_Price tab (AAPL is the current stock).
2. For each derived column, reproduce the computation in Python from the raw close prices.
3. Compare cell-by-cell. Any divergence = bug.

Sheet layout (verified from formulas, header row 5, data starts row 6):
  C = (date string?)
  D = Date
  E = Close price (from GOOGLEFINANCE)
  F = Price_300DMA   formula: LWMA(E, 300) starting row 305
  G = 300DMA_Gain    formula: F[t]/F[t-1] - 1 starting row 306
  H = Price_200DMA   LWMA(E, 200) starting row 205
  I = 200DMA_Gain    H[t]/H[t-1] - 1
  J = Price_100DMA   LWMA(E, 100) starting row 105
  K = 100DMA_Gain    J[t]/J[t-1] - 1
  L = Price_50DMA    LWMA(E, 50)
  M = Price_20DMA    LWMA(E, 20)
  N = Price Gain     E[t]/E[t-1] - 1 starting row 7
  O = Gain_200DMA    LWMA(N, 200)
  P = Gain_200DMA_50DMA  LWMA(O, 50) - baseline
  Q = Gain_100DMA    LWMA(N, 100)
  R = Gain_100DMA_50DMA  LWMA(Q, 50) - $R$4 (baseline)
  S = Price_Signal_50DMA  AVERAGE(T, 50)
  T = Price_Signal   SUM(V:X)/1000
  U = STD_Signal     [broken cross-sheet ref]
  V = Price_GT_50DMA IF(E>L, 1, -1)
  W = 50DMA_GT_100DMA IF(L>J, 1, -1)
  X = 100DMA_GT_200DMA IF(J>H, 1, -1)
  Y = Momentum_100DMA_GT_200DMA  IF(R>P, 1, -1)
  AB = Strategy 1 amount  IF(R<0, -B, 0); last row = -E[last] * AC[last]
  AC3 = COUNTIF(AB<=-1) — count of fires
  AC4 = XIRR(AB6:AB4941, D6:D4941)
  B = weight = 1, 2, 3, ..., N (linear ramp)
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

XLSX = Path(r"c:\Users\vilas\Downloads\AI_Hedge_Fund\extracted\sheets\Stocks_Buy_Signal_Analyser_15Yr.xlsx")
OUT_CSV = Path(r"c:\Users\vilas\Downloads\AI_Hedge_Fund\extracted\_aapl_sheet_data.csv")


def extract_sheet_data() -> pd.DataFrame:
    """Pull every relevant column from the Stock_Price tab as cached values."""
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Stock_Price"]
    # Determine actual data extent
    print(f"Sheet dims: {ws.max_row} rows × {ws.max_column} cols")

    # Column letters → meaningful names (matches header row 5)
    cols = {
        "B": "weight",
        "D": "date",
        "E": "close",
        "F": "MA_300",
        "G": "MA_300_gain",
        "H": "MA_200",
        "I": "MA_200_gain",
        "J": "MA_100",
        "K": "MA_100_gain",
        "L": "MA_50",
        "M": "MA_20",
        "N": "price_gain",
        "O": "gain_200dma",
        "P": "gain_200dma_50dma",
        "Q": "gain_100dma",
        "R": "gain_100dma_50dma",
        "T": "price_signal",
        "V": "price_gt_50",
        "W": "ma50_gt_ma100",
        "X": "ma100_gt_ma200",
        "AB": "strategy1_amt",
    }

    rows = {}
    data_start = 6
    data_end = min(ws.max_row, 5000)
    for col_letter, name in cols.items():
        vals = []
        for r in range(data_start, data_end + 1):
            v = ws[f"{col_letter}{r}"].value
            vals.append(v)
        rows[name] = vals

    df = pd.DataFrame(rows)
    # Cast where sensible
    for c in df.columns:
        if c == "date":
            df[c] = pd.to_datetime(df[c], errors="coerce")
        else:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    # Use date as index
    df = df.dropna(subset=["date"]).set_index("date")
    print(f"Data rows after dropna(date): {len(df)}")
    print(f"Date range: {df.index[0]} -> {df.index[-1]}")
    return df


def main():
    df = extract_sheet_data()
    print()
    print("First 5 rows of each column:")
    print(df.head(5).to_string())
    print()
    print("Sample values at key boundaries (where each LWMA first activates):")
    # Sheet row 105 = MA_100 first value → 0-indexed 99 in our df (since data starts row 6 = idx 0)
    for excel_row in [25, 55, 105, 155, 205, 305]:
        idx = excel_row - 6
        if idx >= len(df):
            continue
        print(f"\n  Excel row {excel_row} (df index {idx}, date {df.index[idx].date()}):")
        for col in ["close", "MA_20", "MA_50", "MA_100", "MA_200", "MA_300",
                    "MA_100_gain", "price_gain", "gain_100dma", "gain_100dma_50dma",
                    "price_signal", "strategy1_amt"]:
            v = df[col].iloc[idx]
            print(f"    {col:<22} = {v}")

    print()
    print("Statistics on Strategy 1 amount column (raw cached values):")
    amt = df["strategy1_amt"]
    nonzero = amt[amt != 0]
    print(f"  Non-zero entries: {len(nonzero)} of {len(amt)}")
    print(f"  Negative entries (buys): {(amt < 0).sum()}")
    print(f"  Positive entries (sells): {(amt > 0).sum()}")
    print(f"  Sum: {amt.sum():,.2f}")
    if (amt > 0).sum() > 0:
        print(f"  Largest positive (closeout?): {amt[amt > 0].iloc[0]:,.2f} on {amt[amt > 0].index[0].date()}")

    # Save the full table
    df.to_csv(OUT_CSV)
    print(f"\nSaved to {OUT_CSV}")


if __name__ == "__main__":
    main()
